import pandas as pd
from openpyxl import load_workbook
import glob
import random
file_url=""
while(1):


    file_url=input(f"请输入excel文件列表地址(例如：D:/桌面/学校文件/计算机235查寝情况)：")

    # 获取所有Excel文件的文件名
    excel_files = glob.glob(file_url+"\*.xlsx")

    # 循环读取每个Excel文件的表格内容
    for file in excel_files:
        #随机数生成
        random_numbers = [random.randint(1, 5) for _ in range(6)]
        max_numbers = [random.randint(10, 10) for _ in range(4)]
        max01_numbers = [random.randint(5, 8) for _ in range(1)]
        max02_numbers = [random.randint(10, 10) for _ in range(2)]
        re=random_numbers + max_numbers + max01_numbers + max02_numbers
        #以上为男生寝室分数代码

        max010_numbers = [random.randint(10, 10) for _ in range(1)]
        all_re=random_numbers + max_numbers + max010_numbers + max02_numbers
        #以上为女生寝室分数代码

        print(f"文件名称为: {file}")
        # 读取Excel文件
        wb = load_workbook(file)
        sheet = wb.active

        if any(num in file for num in ["6514", "6525", "6526", "6527"]):
            re_er = all_re
        else:
            re_er = re
        re_e=iter(re_er)
        print(re_er)
        # for i in range(3, 16):  # 遍历行索引从3到15（不包括15）
        #     user_input = input(f"请输入第{i}行要修改的数据：")
        
        #     try:
        #         user_input = int(user_input)  # 将用户输入的内容转换为浮点数
        #     except ValueError:
        #         print("输入的内容无法转换为数值")
        #         continue
        
        #     sheet.cell(row=i, column=5, value=user_input)  # 将转换后的值赋给指定的单元格
        #re_iter = iter(re)  # 创建 re 的迭代器
        
        for i in range(3, 16):  # 遍历行索引从3到15（不包括15）
            try:
                user_input =next(re_e)  # 从迭代器中取出下一个元素
                user_input = int(user_input)  # 将用户输入的内容转换为浮点数
            except ValueError:
                print("输入的内容无法转换为数值")
                continue
            sheet.cell(row=i, column=5, value=user_input)  # 将转换后的值赋给指定的单元格
        wb.save(file)
        print(f"处理后的数据已保存至 {file}")